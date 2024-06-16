from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from .models import Archivo, Grafica
from .forms import ArchivoForm
from django.http import FileResponse
from django.conf import settings
import os
import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from django.contrib.auth.views import LoginView

# Create your views here.

def home(request):
    return render(request, 'accounts/home.html')

def developer(request):
    return render(request, 'accounts/developer.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('welcome')
        else:
            return render(request, 'accounts/login.html', {'error': 'Invalid credentials'})
    return render(request, 'accounts/login.html')

def logout_view(request):
    logout(request)
    return redirect('home')

def register(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        if User.objects.filter(username=username).exists():
            return render(request, 'accounts/register.html', {'error': 'Username already exists'})
        user = User.objects.create_user(username=username, password=password)
        login(request, user)
        return redirect('welcome')
    return render(request, 'accounts/register.html')

@login_required
def welcome(request):
    if request.method == 'POST':
        form = ArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.save(commit=False)
            archivo.usuario = request.user
            archivo.save()
            return redirect('welcome')
    else:
        form = ArchivoForm()
    archivos = Archivo.objects.filter(usuario=request.user)
    return render(request, 'accounts/welcome.html', {'form': form, 'archivos': archivos})

@login_required
def serve_file(request, archivo_id):
    archivo = Archivo.objects.get(id=archivo_id)
    file_path = os.path.join(settings.MEDIA_ROOT, archivo.archivo.name)
    return FileResponse(open(file_path, 'rb'))

@login_required
def generate_chart(request, archivo_id):
    archivo = get_object_or_404(Archivo, id=archivo_id)
    chart_type = request.POST.get('chart_type')
    
    # Ruta al archivo XLSX en el sistema de archivos
    file_path = archivo.archivo.path
    
    # Abrir el archivo XLSX
    wb = openpyxl.load_workbook(file_path)
    
    # Seleccionar la primera hoja de cálculo (puedes ajustar esto según tu archivo)
    sheet = wb.active
    
    # Leer los datos del archivo y convertirlos en una lista
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    # Procesar los datos para la gráfica
    labels = [str(row[0]) for row in data[1:]]  # Suponiendo que las etiquetas están en la primera columna
    values = [row[1] for row in data[1:]]  # Suponiendo que los valores están en la segunda columna
    
    if chart_type == 'bar':
        # Generar la gráfica de barras
        plt.figure(figsize=(10, 6))
        plt.bar(np.arange(len(labels)), values)
        plt.xticks(np.arange(len(labels)), labels, rotation=45)
        plt.xlabel('Category')
        plt.ylabel('Value')
        plt.title('Bar Chart')
        chart_file_name = f'grafica_bar_{archivo_id}.png'
    elif chart_type == 'pie':
        # Generar la gráfica de pastel
        plt.figure(figsize=(10, 6))
        plt.pie(values, labels=labels, autopct='%1.1f%%', startangle=140)
        plt.title('Pie Chart')
        chart_file_name = f'grafica_pie_{archivo_id}.png'
    elif chart_type == 'line':
        # Generar la gráfica de líneas
        plt.figure(figsize=(10, 6))
        plt.plot(labels, values, marker='o')
        plt.xticks(rotation=45)
        plt.xlabel('Category')
        plt.ylabel('Value')
        plt.title('Line Chart')
        chart_file_name = f'grafica_line_{archivo_id}.png'
    else:
        return redirect('archivo_detail', archivo_id=archivo_id)
    
    chart_file_path = os.path.join(settings.MEDIA_ROOT, chart_file_name)
    plt.savefig(chart_file_path)
    plt.close()
    
    # Crear una instancia de Grafica y guardarla en la base de datos
    Grafica.objects.create(archivo=archivo, tipo_grafica=chart_type, parametros={'chart_file_path': os.path.join(settings.MEDIA_URL, chart_file_name)})
    
    # Redirigir a la página de detalles del archivo
    return redirect('archivo_detail', archivo_id=archivo_id)

@login_required
def generate_pie_chart(request, archivo_id):
    archivo = get_object_or_404(Archivo, id=archivo_id)
    return generate_chart(request, archivo_id)

@login_required
def serve_file(request, archivo_id):
    archivo = Archivo.objects.get(id=archivo_id)
    file_path = os.path.join(settings.MEDIA_ROOT, archivo.archivo.name)
    return FileResponse(open(file_path, 'rb'))

@login_required
def archivo_detail(request, archivo_id):
    archivo = get_object_or_404(Archivo, id=archivo_id)
        
    # Leer el archivo XLSX y obtener los datos
    file_path = archivo.archivo.path
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Convertir los datos de la hoja de cálculo en una lista
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Obtener todas las gráficas asociadas
    graficas = Grafica.objects.filter(archivo=archivo)
    
    # Verificar si ya existe una gráfica de barras, de pastel o de líneas
    bar_chart_generated = False
    pie_chart_generated = False
    line_chart_generated = False
    for grafica in graficas:
        if grafica.tipo_grafica == 'bar':
            bar_chart_generated = True
        elif grafica.tipo_grafica == 'pie':
            pie_chart_generated = True
        elif grafica.tipo_grafica == 'line':
            line_chart_generated = True

    context = {
        'archivo': archivo,
        'data': data,
        'graficas': graficas,
        'bar_chart_generated': bar_chart_generated,
        'pie_chart_generated': pie_chart_generated,
        'line_chart_generated': line_chart_generated,
    }
    return render(request, 'accounts/archivo_detail.html', context)

@login_required
def delete_archivo(request, archivo_id):
    archivo = get_object_or_404(Archivo, id=archivo_id)
    archivo.delete()
    return redirect('welcome')
